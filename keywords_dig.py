from playwright.sync_api import sync_playwright
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os
from urllib.parse import urlparse
import requests
import argparse  # 添加到文件开头的导入部分
import json
import uuid

# 飞书API配置
APP_ID = "cli_a7eab65ecc98501c"
APP_SECRET = "DIiBXmld4IruCTE6smLvFhYaTqbi5bDH"
APP_TOKEN = "KB8tbjtK4antdNsp26TcEGabnXe"

class TokenManager:
    def __init__(self):
        self._token = None
        self._token_expire_time = None
    
    def get_token(self):
        """获取有效的tenant_access_token"""
        current_time = datetime.now()
        
        # 如果没有token或者token已经过期或即将过期（小于30分钟），则重新获取
        if (self._token is None or 
            self._token_expire_time is None or 
            current_time + timedelta(minutes=30) >= self._token_expire_time):
            self._refresh_token()
        
        return self._token
    
    def _refresh_token(self):
        """刷新tenant_access_token"""
        url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
        headers = {
            "Content-Type": "application/json; charset=utf-8"
        }
        data = {
            "app_id": APP_ID,
            "app_secret": APP_SECRET
        }
        
        try:
            response = requests.post(url, headers=headers, json=data)
            response_json = response.json()
            
            if response_json.get("code") == 0:
                self._token = response_json.get("tenant_access_token")
                expire_seconds = response_json.get("expire", 7200)
                self._token_expire_time = datetime.now() + timedelta(seconds=expire_seconds)
                print("Token已更新，有效期至:", self._token_expire_time)
            else:
                print(f"获取token失败: {response_json.get('msg')}")
                self._token = None
                self._token_expire_time = None
        except Exception as e:
            print(f"刷新token时发生错误: {str(e)}")
            self._token = None
            self._token_expire_time = None

# 创建全局的token管理器实例
token_manager = TokenManager()

class SemrushCrawler:
    def __init__(self):
        self.url = "https://dash.3ue.com/zh-Hans/#/page/m/home"
        self.username = "belulo"
        self.password = "belulo123456"
        self.download_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
        self.app_token = "OgmkbZFfRao9yjsc7S5c6gw5nzh"
        self.receive_ids = ["bededf32", "ceeb89d9", "4gac44cg"]
        
    def get_tenant_access_token(self):
        """获取访问令牌"""
        try:
            # 获取token
            token = token_manager.get_token()
            if not token:
                raise Exception("获取访问令牌失败")
            return token
        except Exception as e:
            print(f"获取访问令牌失败: {str(e)}")
            return None
    
    def get_competitor_urls_from_feishu(self):
        """从飞书多维表格获取竞品网址列表"""
        try:
            # 获取访问令牌
            token = self.get_tenant_access_token()
            if not token:
                raise Exception("获取访问令牌失败")

            # 构建请求URL和Headers
            base_url = "https://open.feishu.cn/open-apis/bitable/v1/apps"
            url = f"{base_url}/OgmkbZFfRao9yjsc7S5c6gw5nzh/tables/tbljRMQrnSSWV8Ds/records"
            
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json; charset=utf-8"
            }

            print(f"\n请求URL: {url}")
            print("请求Headers:", headers)

            # 发送请求获取数据
            response = requests.get(url, headers=headers)
            print(f"响应状态码: {response.status_code}")
            response_json = response.json()
            print("响应内容:", response_json)

            if response_json.get("code") == 0:
                records = response_json.get("data", {}).get("items", [])
                urls = []
                print("\n开��解析数据...")
                for record in records:
                    # 从记录中提取链接字段
                    link_field = record.get("fields", {}).get("链接")
                    if link_field and isinstance(link_field, dict):
                        url = link_field.get("link", "").strip().rstrip('/')
                        if url:
                            urls.append(url)
                            print(f"找到URL: {url}")

                print(f"\n从飞书获取到 {len(urls)} 个竞品网址:")
                for i, url in enumerate(urls, 1):
                    print(f"{i}. {url}")
                
                return urls
            else:
                raise Exception(f"获取表格数据失败: {response_json.get('msg')}\n错误详�����信息: {response_json.get('error', {})}")

        except Exception as e:
            print(f"\n从飞书获取竞品网址时发生错误: {str(e)}")
            print("错误详信息:", e.__class__.__name__)
            import traceback
            print("错误堆栈:", traceback.format_exc())
            # 如果从飞书获取失败，回退到手动输入
            print("\n从飞书获取失败，切换到手动输入模式...")
            return self.get_competitor_urls_manual()
    
    def get_competitor_urls_manual(self):
        """手动输入竞品网址（作为备用方案）"""
        print("请输入竞品网址(每行一个，输入空行结束):")
        urls = []
        while True:
            url = input().strip()
            if not url:
                break
            urls.append(url.rstrip('/'))
        return urls

    def create_download_directory(self):
        """创建下载目录"""
        date_str = datetime.now().strftime('%Y%m%d')
        folder_name = f"竞品监控_{date_str}"
        download_path = os.path.join(self.download_dir, folder_name)
        
        if not os.path.exists(download_path):
            os.makedirs(download_path)
        
        return download_path
    
    def get_domain_name(self, url):
        """获取URL的������������"""
        parsed_uri = urlparse(url)
        domain = parsed_uri.netloc
        # 确保域名以 http:// 或 https:// 开头
        if not domain:
            domain = urlparse(f"http://{url}").netloc
        # 获取一级域名
        parts = domain.split('.')
        if len(parts) > 2:
            return '.'.join(parts[-2:])
        return domain

    def select_time_period(self, page):
        """选择时间段"""
        try:
            print("正在选择时间段...")
            
            # 等待时间选择器出现并点击
            time_selector = page.wait_for_selector('[data-at="date-selector"]')
            if not time_selector:
                raise Exception("未找到时间选择器")
            
            time_selector.click()
            time.sleep(1)  # 等待下拉菜单出现
            
            # 选择第一个选项（最新时间）
            first_option = page.wait_for_selector('div[data-at="date-option"]:first-child')
            if not first_option:
                raise Exception("未找到时间选项")
            
            first_option.click()
            
            # 等待页面更新
            page.wait_for_load_state('networkidle')
            time.sleep(2)
            
            print("时间段选��完成")
            
        except Exception as e:
            print(f"选择时间段时发生错误: {str(e)}")
            raise

    def click_view_details(self, page):
        """点击查看详情按钮"""
        try:
            print("查找'查看详情'按钮...")
            
            # 等待关键词板块加载
            keywords_section = page.wait_for_selector('section[aria-label="主要自然搜索关键词"]')
            if not keywords_section:
                raise Exception("未找到关键词板块")
            
            # 查找并点击查看详情按钮
            view_details = keywords_section.wait_for_selector('button:has-text("查看详情")')
            if not view_details:
                raise Exception("未找到查看详情按钮")
            
            view_details.click()
            
            # 等待页面加载
            page.wait_for_load_state('networkidle')
            time.sleep(3)
            
            print("已点击查看详情")
            
        except Exception as e:
            print(f"点击查看详情时发生错误: {str(e)}")
            raise

    def create_feishu_table(self, domain_name):
        """创建飞书数据表"""
        try:
            print("\n正在创建飞书数据表...")
            
            # ���建数���表名称
            date_str = datetime.now().strftime('%Y%m%d')
            table_name = f"{date_str}_{domain_name}"
            
            # 构建请求URL和Headers
            url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{self.app_token}/tables"
            headers = {
                "Authorization": f"Bearer {token_manager.get_token()}",
                "Content-Type": "application/json; charset=utf-8"
            }
            
            # 构建请求体
            payload = {
                "table": {
                    "name": table_name,
                    "default_view_name": "关键词数据",
                    "fields": [
                        {
                            "field_name": "关键词",
                            "type": 1  # 文本类型
                        },
                        {
                            "field_name": "KD%",
                            "type": 1
                        },
                        {
                            "field_name": "URL",
                            "type": 1
                        }
                    ]
                }
            }
            
            print(f"创建数据表请求URL: {url}")
            print("请求Headers:", headers)
            print("���求Body:", payload)
            
            # 发送请求
            response = requests.post(url, headers=headers, json=payload)
            print(f"响应状态码: {response.status_code}")
            response_json = response.json()
            print("响应内容:", response_json)
            
            if response_json.get("code") == 0:
                table_id = response_json["data"]["table_id"]
                print(f"数据表创建成功，table_id: {table_id}")
                return table_id
            else:
                raise Exception(f"创建数据表失败: {response_json.get('msg')}")
                
        except Exception as e:
            print(f"创建飞书数据表时发生错误: {str(e)}")
            return None
    
    def upload_data_to_feishu(self, data, table_id):
        """上传数据到飞书数据表"""
        try:
            print("\n正在上传数据到飞书...")
            
            # 构建请求URL和Headers
            url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/batch_create"
            headers = {
                "Authorization": f"Bearer {token_manager.get_token()}",
                "Content-Type": "application/json; charset=utf-8"
            }
            
            # 构建记录数据
            records = []
            for item in data:
                record = {
                    "fields": {
                        "关键词": item["keyword"],
                        "KD%": item["kd"],
                        "URL": item["url"]
                    }
                }
                records.append(record)
            
            # 分批上传数据（每次最多100条）
            batch_size = 100
            for i in range(0, len(records), batch_size):
                batch = records[i:i + batch_size]
                payload = {"records": batch}
                
                print(f"\n正在上传第 {i//batch_size + 1} 批数据...")
                response = requests.post(url, headers=headers, json=payload)
                response_json = response.json()
                
                if response_json.get("code") != 0:
                    raise Exception(f"上传数据失败: {response_json.get('msg')}")
                
                print(f"成功上传 {len(batch)} 条记录")
            
            print(f"\n所有数据上传完成，共 {len(records)} 条记录")
            return True
            
        except Exception as e:
            print(f"上传数据到飞书时发生错误: {str(e)}")
            return False

    def search_competitor(self, page, competitor_url):
        """搜索竞品网址并抓取数据"""
        try:
            print(f"正在搜索竞品网址: {competitor_url}")
            
            # 重新加载项目页面
            print("重新加载项目页面...")
            page.goto('https://sem.3ue.com/projects/?__gmitm=aHR0cHM6Ly96aC5zZW1ydXNoLmNvbQ--', wait_until='networkidle')
            time.sleep(3)  # 等待页面完全加载
            
            # 确保页面已经完全加载
            page.wait_for_load_state('networkidle')
            print("等��搜索框出现...")
            
            # 先尝试等待搜索框的父容器
            page.wait_for_selector('#srf-search-bar', state='visible', timeout=30000)
            
            # 然后等待搜索框
            search_input = page.wait_for_selector(
                '[data-test="searchbar_input"]',
                state='visible',
                timeout=30000
            )
            
            if not search_input:
                raise Exception("未找到搜索框")
            
            # 确保页面稳定后再操作
            time.sleep(2)
            
            # 清空搜索框并输入竞品网址
            search_input.click()
            page.wait_for_timeout(500)
            search_input.fill(competitor_url)
            
            # 使用更稳定的搜索按钮选择器
            search_button = page.wait_for_selector(
                '[data-test="searchbar_search_submit"]',
                state='visible',
                timeout=30000
            )
            search_button.click()
            
            # 等待搜索结果加载
            page.wait_for_load_state('networkidle')
            time.sleep(3)
            
            print("搜索完成")
            
            # 选择时间段
            self.select_time_period(page)
            
            # 点击查看详情
            self.click_view_details(page)
            
            # 点击自然搜索
            self.click_organic_search(page)
            
            # 抓取数据
            data = self.extract_table_data(page)
            
            print(f"\n总共抓取到 {len(data)} 条数据")
            
            # 保存到Excel
            self.save_to_excel(data, competitor_url)
            
            # 创建飞书数据表并上传数据
            domain_name = self.get_domain_name(competitor_url)
            table_id = self.create_feishu_table(domain_name)
            if table_id:
                self.upload_data_to_feishu(data, table_id)
            
            return data
            
        except Exception as e:
            print(f"搜索竞品时发生错误: {str(e)}")
            raise
    
    def login(self, page, username, password):
        """登录功能"""
        try:
            print("正在尝试登录...")
            
            # 等待用户名输入框出现
            username_input = page.wait_for_selector('#input-username')
            if not username_input:
                raise Exception("未找到用户名输入框")
                
            # 等待密码输入框出现
            password_input = page.wait_for_selector('#input-password')
            if not password_input:
                raise Exception("未��到密码输入框")
            
            # 输入用名和密码
            username_input.fill(username)
            password_input.fill(password)
            
            # 点击登录按钮
            login_button = page.wait_for_selector('button:has-text("登录")')
            if not login_button:
                raise Exception("未找到登录按钮")
                
            login_button.click()
            
            # 等待登录完成
            page.wait_for_load_state('networkidle')
            time.sleep(3)  # 额外等待确保登录成功
            
            print("登录完成")
            
        except Exception as e:
            print(f"登录过程中发生错误: {str(e)}")
            raise
    
    def start(self):
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(viewport={'width': 1920, 'height': 1080})
            page = context.new_page()
            
            try:
                # 设置页面加载超时时间
                page.set_default_timeout(60000)
                
                print("正在打开SEMrush...")
                page.goto(self.url)
                page.wait_for_load_state('networkidle')
                
                # 使用预设的登录��据
                print(f"正在使用账号 {self.username} 登录...")
                self.login(page, self.username, self.password)
                
                # 从飞书获取竞品网址
                competitor_urls = self.get_competitor_urls_from_feishu()
                if not competitor_urls:
                    print("未获取到任何网址，程序退出")
                    return
                
                print(f"\n共获取到 {len(competitor_urls)} 个竞品网址")
                
                # 主动打开新标签页
                print("正在打开项目页面...")
                new_page = context.new_page()
                project_url = 'https://sem.3ue.com/projects/?__gmitm=aHR0cHM6Ly96aC5zZW1ydXNoLmNvbQ--'
                
                try:
                    new_page.goto(project_url, wait_until='networkidle')
                    
                    # 循环处理每个竞品网址
                    for i, competitor_url in enumerate(competitor_urls, 1):
                        print(f"\n正在处理第 {i}/{len(competitor_urls)} 个网址: {competitor_url}")
                        self.search_competitor(new_page, competitor_url)
                        
                        # 如果不是最后一个���址，等待一段时间再处理下一个
                        if i < len(competitor_urls):
                            print("等待5秒后处理下一个网址...")
                            time.sleep(5)
                    
                except Exception as e:
                    print(f"处理竞品网址时发生错误: {str(e)}")
                    raise
                    
                # 在所有竞品处理完成后，进行关键词比较
                print("\n所有竞品处理完成，开始分析新增关键词...")
                self.compare_keywords_data()
                
            except Exception as e:
                print(f"发生错误: {str(e)}")
            finally:
                context.close()
                browser.close()

    def click_organic_search(self, page):
        """点击自然搜索选项"""
        try:
            print("正在切换到自然搜索...")
            organic_search = page.wait_for_selector('div[data-ui-name="Box"] div[data-ui-name="Ellipsis"]:has-text("自然搜索")')
            if not organic_search:
                raise Exception("未找到'自然搜索'选项")
            organic_search.click()
            
            # 等待表格加载
            page.wait_for_selector('section[aria-label="自然搜索排名"]')
            print("已切换到自然搜索视图")
            
        except Exception as e:
            print(f"切换到自然搜索时发生错误: {str(e)}")
            raise

    def extract_table_data(self, page):
        """抓取表格数据"""
        try:
            data = []
            
            # 抓取前3页数据
            for page_num in range(3):
                print(f"正在抓取第 {page_num + 1} 页数据...")
                
                # 等待表格载完成
                page.wait_for_selector('div[role="table"]')
                time.sleep(2)
                
                # 获取所有行
                rows = page.query_selector_all('h3[data-at="table-row"]')
                
                for row in rows:
                    try:
                        # 抓取关键词
                        keyword_element = row.query_selector('span.___SText_pr68d-red-team[data-ui-name="Link.Text"]')
                        keyword = keyword_element.inner_text() if keyword_element else "N/A"
                        
                        # 抓取KD%
                        kd_element = row.query_selector('div[data-at="display-kd"] span[data-at="kd-value"]')
                        kd_value = kd_element.inner_text() if kd_element else "不可用"
                        
                        # 抓���URL - 更新选择器
                        url_element = row.query_selector('div[data-ui-name="Link.Text"].___SEllipsis_s2vlm-red-team')
                        url = url_element.inner_text() if url_element else "N/A"
                        
                        data.append({
                            'keyword': keyword,
                            'kd': kd_value,
                            'url': url
                        })
                    except Exception as e:
                        print(f"处理行数据时发生错误: {str(e)}")
                        continue
                
                # 打印当前页数据
                if len(data) > 0:
                    print("\n当前页数据:")
                    print("{:<50} {:<10} {:<100}".format("关键词", "KD%", "URL"))
                    print("-" * 160)
                    for item in data[-len(rows):]:
                        print("{:<50} {:<10} {:<100}".format(
                            item['keyword'][:50],
                            item['kd'],
                            item['url'][:100]
                        ))
                else:
                    print("当前页未找到数据")
                
                # 点击下一页
                if page_num < 2:  # 只在前2页点击下一页
                    next_button = page.query_selector('button[data-at="next-page"]:not([disabled])')
                    if next_button:
                        next_button.click()
                        page.wait_for_load_state('networkidle')
                        time.sleep(3)  # 增加等待时间
                    else:
                        print("没有更多页面了")
                        break
                        
            return data
            
        except Exception as e:
            print(f"抓取数据时发生错误: {str(e)}")
            raise

    def save_to_excel(self, data, competitor_url):
        """保存数据到Excel"""
        try:
            # 建下载目录
            download_path = self.create_download_directory()
            
            # 获取域名和日期用文件命名
            domain_name = self.get_domain_name(competitor_url)
            date_str = datetime.now().strftime('%Y%m%d')
            filename = f"{domain_name}_{date_str}.xlsx"
            filepath = os.path.join(download_path, filename)
            
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = "�����词数据"
            
            # 设置列宽
            ws.column_dimensions['A'].width = 50  # 关键词
            ws.column_dimensions['B'].width = 15  # KD%
            ws.column_dimensions['C'].width = 80  # URL
            
            # 设置标题行样式
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name='微软雅黑', size=12, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入标题
            headers = ['关键词', 'KD%', 'URL']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 设置数据行样式
            data_font = Font(name='微软雅黑', size=11)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # 写入数据
            for row, item in enumerate(data, 2):
                # 设置行高
                ws.row_dimensions[row].height = 25
                
                # 写入并设置单元格样式
                cells = [
                    (item['keyword'], 'A'),
                    (item['kd'], 'B'),
                    (item['url'], 'C')
                ]
                
                for value, col in cells:
                    cell = ws[f'{col}{row}']
                    cell.value = value
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    
                    # 为偶数行添加浅色背景
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            # 添加统计信息
            summary_row = len(data) + 3
            ws[f'A{summary_row}'] = f"竞品网址: {competitor_url}"
            ws[f'A{summary_row+1}'] = f"抓取时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{summary_row+2}'] = f"总计关键词数: {len(data)}"
            
            # ���并统计信息单元格
            for i in range(3):
                ws.merge_cells(f'A{summary_row+i}:C{summary_row+i}')
                cell = ws[f'A{summary_row+i}']
                cell.font = Font(name='微软雅黑', size=11, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 保存文件
            wb.save(filepath)
            print(f"\n数据已保存到: {filepath}")
            
        except Exception as e:
            print(f"保存Excel时发生错误: {str(e)}")
            raise

    def get_all_tables(self):
        """获取所有数据表信息"""
        try:
            print("\n正在获取所有数据表信息...")
            url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{self.app_token}/tables"
            headers = {
                "Authorization": f"Bearer {token_manager.get_token()}",
                "Content-Type": "application/json; charset=utf-8"
            }
            
            response = requests.get(url, headers=headers)
            response_json = response.json()
            
            if response_json.get("code") == 0:
                tables = response_json["data"]["items"]
                # 筛选符合"日期_域"格式的表
                keyword_tables = []
                for table in tables:
                    name = table["name"]
                    if self.is_keyword_table(name):
                        keyword_tables.append({
                            "name": name,
                            "table_id": table["table_id"],
                            "date": name.split('_')[0],
                            "domain": name.split('_')[1]
                        })
                return keyword_tables
            else:
                raise Exception(f"获取数据表失败: {response_json.get('msg')}")
        except Exception as e:
            print(f"获取数据表信息时发生错误: {str(e)}")
            return []

    def is_keyword_table(self, table_name):
        """判断是否为关键词数据表（日期_域名格式）"""
        try:
            parts = table_name.split('_')
            if len(parts) != 2:
                return False
            date_str, domain = parts
            # 验证日期格式（YYYYMMDD）
            datetime.strptime(date_str, '%Y%m%d')
            return True
        except:
            return False

    def get_table_data(self, table_id):
        """获取指定数据表的所有记录"""
        try:
            url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{self.app_token}/tables/{table_id}/records/search"
            headers = {
                "Authorization": f"Bearer {token_manager.get_token()}",
                "Content-Type": "application/json; charset=utf-8"
            }
            
            # 构建查询条件
            payload = {
                "page_size": 100,  # 每页记录数
                "field_names": ["关键词", "KD%", "URL"]
            }
            
            all_records = []
            while True:
                response = requests.post(url, headers=headers, json=payload)
                response_json = response.json()
                
                if response_json.get("code") == 0:
                    records = response_json["data"]["items"]
                    all_records.extend(records)
                    
                    # 检查是否还有更多数据
                    if not response_json["data"].get("has_more"):
                        break
                    
                    # 更新分页标记
                    payload["page_token"] = response_json["data"].get("page_token")
                else:
                    raise Exception(f"获取表格数据失败: {response_json.get('msg')}")
            
            return all_records
        except Exception as e:
            print(f"获取表格数据时���生错误: {str(e)}")
            return []

    def send_feishu_message(self, message_content):
        """发送飞书消息到多个接收者"""
        success = True
        for receive_id in self.receive_ids:
            try:
                print(f"\n正在发送飞书消息到用户 {receive_id}...")
                url = "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=user_id"
                headers = {
                    "Authorization": f"Bearer {token_manager.get_token()}",
                    "Content-Type": "application/json"
                }
                
                # 生成UUID
                message_uuid = str(uuid.uuid4())
                
                # 构建消息内容
                payload = json.dumps({
                    "receive_id": receive_id,
                    "msg_type": "text",
                    "content": json.dumps({"text": message_content}),
                    "uuid": message_uuid
                })
                
                print("发送消息请求:")
                print(f"URL: {url}")
                print(f"Headers: {headers}")
                print(f"Payload: {payload}")
                
                response = requests.post(url, headers=headers, data=payload)
                response_json = response.json()
                
                print("响应内容:", response_json)
                
                if response_json.get("code") == 0:
                    print(f"消息成功发送到用户 {receive_id}")
                else:
                    print(f"发送消息到用户 {receive_id} 失败: {response_json.get('msg')}")
                    success = False
                    
            except Exception as e:
                print(f"发送消息到用户 {receive_id} 时发生错误: {str(e)}")
                success = False
                continue  # 继续发送给其他用户
        
        return success

    def format_keywords_message(self, domain, new_keywords_data, latest_date, previous_date):
        """格式化关键词消息内容"""
        message_lines = [
            f"域名 {domain} 新增关键词通知",
            f"对比时间: {latest_date} vs {previous_date}",
            f"新增关键词数量: {len(new_keywords_data)}",
            "\n关键词详情:"
        ]
        
        for data in new_keywords_data:
            message_lines.append(
                f"关键词: {data['keyword']}\n"
                f"KD%: {data['kd']}\n"
                f"URL: {data['url']}\n"
            )
        
        return "\n".join(message_lines)

    def compare_keywords_data(self):
        """比较关键词数据并找出新增内容"""
        try:
            print("\n开始比较关键词数据...")
            
            # 获取所有关键词数据表
            tables = self.get_all_tables()
            if not tables:
                print("未找到关键词数据表")
                return
            
            # 按域名分组
            domain_tables = {}
            for table in tables:
                domain = table["domain"]
                if domain not in domain_tables:
                    domain_tables[domain] = []
                domain_tables[domain].append(table)
            
            # 对每个域名进行数据比较
            for domain, tables in domain_tables.items():
                print(f"\n处理域名: {domain}")
                
                # 按日期排序
                tables.sort(key=lambda x: x["date"], reverse=True)
                
                if len(tables) < 2:
                    print(f"域名 {domain} 只有一个数据表，无法比较")
                    continue
                
                # 获取最新和次新的数据
                latest_table = tables[0]
                previous_table = tables[1]
                
                # 获取对比的日期
                latest_date = latest_table["date"]
                previous_date = previous_table["date"]
                
                print(f"比较数据表: {latest_table['name']} vs {previous_table['name']}")
                
                # 获取两个表的数据
                latest_data = self.get_table_data(latest_table["table_id"])
                previous_data = self.get_table_data(previous_table["table_id"])
                
                # 打印原始数据结构（用于调试）
                print("\n最新数据示例:")
                if latest_data:
                    print(latest_data[0])
                
                # 转换为集合以便比较
                latest_keywords = set()
                previous_keywords = set()
                
                # 处理最新数据
                for record in latest_data:
                    keyword_field = record["fields"].get("关键词", [])
                    if keyword_field and isinstance(keyword_field, list):
                        keyword = keyword_field[0].get("text", "") if keyword_field[0] else ""
                        if keyword:
                            latest_keywords.add(keyword)
                
                # 处理历史数据
                for record in previous_data:
                    keyword_field = record["fields"].get("关键词", [])
                    if keyword_field and isinstance(keyword_field, list):
                        keyword = keyword_field[0].get("text", "") if keyword_field[0] else ""
                        if keyword:
                            previous_keywords.add(keyword)
                
                # 找出新增的关键词
                new_keywords = latest_keywords - previous_keywords
                
                if new_keywords:
                    print(f"\n域名 {domain} 新增 {len(new_keywords)} 个关键词:")
                    print("{:<50} {:<10} {:<100}".format("关键词", "KD%", "URL"))
                    print("-" * 160)
                    
                    # 收集新增关键词的详细信息
                    new_keywords_data = []
                    
                    # 打印新增的关键词详细信息
                    for record in latest_data:
                        keyword_field = record["fields"].get("关键词", [])
                        if keyword_field and isinstance(keyword_field, list):
                            keyword = keyword_field[0].get("text", "") if keyword_field[0] else ""
                            
                            if keyword in new_keywords:
                                kd_field = record["fields"].get("KD%", [])
                                kd = kd_field[0].get("text", "N/A") if kd_field and kd_field[0] else "N/A"
                                
                                url_field = record["fields"].get("URL", [])
                                url = url_field[0].get("text", "N/A") if url_field and url_field[0] else "N/A"
                                
                                # 添加到新增关键词数据列表
                                new_keywords_data.append({
                                    'keyword': keyword,
                                    'kd': kd,
                                    'url': url
                                })
                                
                                print("{:<50} {:<10} {:<100}".format(
                                    str(keyword)[:50],
                                    str(kd),
                                    str(url)[:100]
                                ))
                    
                    # 发送飞书通知
                    if new_keywords_data:
                        message_content = self.format_keywords_message(
                            domain, 
                            new_keywords_data,
                            latest_date,
                            previous_date
                        )
                        self.send_feishu_message(message_content)
                else:
                    print(f"域名 {domain} 没有新增关键词")
            
        except Exception as e:
            print(f"比较关键词数据时发生错误: {str(e)}")
            import traceback
            print("错误堆栈:", traceback.format_exc())

if __name__ == "__main__":
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='SEMrush关键词挖掘工��')
    parser.add_argument('--compare-only', action='store_true', 
                        help='只执行关键词比对功能，不进行���据抓取')
    args = parser.parse_args()

    crawler = SemrushCrawler()
    
    if args.compare_only:
        # 只执行关键词比对
        print("执行关键词比对...")
        crawler.compare_keywords_data()
    else:
        # 执行完整流程
        crawler.start()
